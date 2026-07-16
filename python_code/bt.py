import tkinter as tk
from tkinter import messagebox, scrolledtext
import imaplib
import email
from email.header import decode_header
import os
import time
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

# =====================================================================
# CẤU HÌNH HỆ THỐNG
# =====================================================================
FILE_EXCEL = "Danh_Sach_Email_Da_Doc.xlsx"
THOI_GIAN_QUET_GIAY = 10  # Cứ 10 giây sẽ kiểm tra hộp thư 1 lần

# Biến toàn cục để điều khiển luồng chạy ngầm
is_monitoring = False
monitor_thread = None


def giai_ma_tieu_de(header_val):
    """Hàm hỗ trợ dịch ngôn ngữ (tiếng Việt có dấu) trong tiêu đề Email"""
    if not header_val:
        return ""
    decoded_list = decode_header(header_val)
    header_str = ""
    for content, encoding in decoded_list:
        if isinstance(content, bytes):
            try:
                header_str += content.decode(encoding or "utf-8", errors="ignore")
            except Exception:
                header_str += content.decode("utf-8", errors="ignore")
        else:
            header_str += str(content)
    return header_str


def lay_noi_dung_email(msg):
    """Hàm trích xuất nội dung văn bản (Body) của email"""
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))
            if content_type == "text/plain" and "attachment" not in content_disposition:
                try:
                    body = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                    break
                except Exception:
                    pass
    else:
        try:
            body = msg.get_payload(decode=True).decode("utf-8", errors="ignore")
        except Exception:
            pass
    return body.strip()


def ghi_vao_excel(mail_data):
    """Hàm ghi tiếp 1 dòng dữ liệu mới vào file Excel duy nhất"""
    if not os.path.exists(FILE_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Email Đã Đọc"
        ws.append(["Thời gian đọc", "Người gửi (From)", "Tiêu đề (Subject)", "Nội dung (Body)"])
    else:
        wb = load_workbook(FILE_EXCEL)
        ws = wb.active

    ws.append([
        mail_data["thoi_gian"],
        mail_data["nguoi_gui"],
        mail_data["tieu_de"],
        mail_data["noi_dung"]
    ])
    wb.save(FILE_EXCEL)


def them_log(noidung):
    """Hiển thị thông báo lên bảng nhật ký trên giao diện"""
    txt_log.config(state="normal")
    thoi_gian = datetime.now().strftime("%H:%M:%S")
    txt_log.insert(tk.END, f"[{thoi_gian}] {noidung}\n")
    txt_log.see(tk.END)  # Tự động cuộn xuống dòng mới nhất
    txt_log.config(state="disabled")


def luong_theo_doi_ngam(user_email, app_password):
    """Luồng xử lý ngầm: kiểm tra email mỗi 10 giây"""
    global is_monitoring
    mail = None
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        mail.login(user_email, app_password)
        mail.select("inbox")
        
        # BƯỚC 1: Quét danh sách các email ĐÃ ĐỌC cũ (Trước khi chạy app)
        _, messages = mail.search(None, "SEEN")
        id_da_doc_cu = set(messages[0].split())
        
        them_log(f"Đã kết nối! Bỏ qua {len(id_da_doc_cu)} email đã đọc trước đây.")
        them_log("Đang bắt đầu theo dõi email mới được đọc...")
        
        # BƯỚC 2: Vòng lặp kiểm tra liên tục khi app đang bật
        while is_monitoring:
            time.sleep(THOI_GIAN_QUET_GIAY)
            if not is_monitoring:
                break
                
            try:
                # Quét lại danh sách email đã đọc hiện tại
                mail.recent()
                _, messages = mail.search(None, "SEEN")
                id_da_doc_hien_tai = set(messages[0].split())
                
                # Tìm ra những email MỚI chuyển sang trạng thái đã đọc
                id_moi_doc = id_da_doc_hien_tai - id_da_doc_cu
                
                if id_moi_doc:
                    for e_id in id_moi_doc:
                        _, msg_data = mail.fetch(e_id, "(RFC822)")
                        for response_part in msg_data:
                            if isinstance(response_part, tuple):
                                msg = email.message_from_bytes(response_part[1])
                                
                                nguoi_gui = giai_ma_tieu_de(msg.get("From"))
                                tieu_de = giai_ma_tieu_de(msg.get("Subject"))
                                noi_dung = lay_noi_dung_email(msg)
                                thoi_gian = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                
                                mail_data = {
                                    "thoi_gian": thoi_gian,
                                    "nguoi_gui": nguoi_gui,
                                    "tieu_de": tieu_de,
                                    "noi_dung": noi_dung
                                }
                                
                                # Ghi ngay vào Excel
                                ghi_vao_excel(mail_data)
                                them_log(f"-> ĐÃ XUẤT EXCEL: {tieu_de[:30]}...")
                    
                    # Cập nhật lại danh sách gốc để không tải lại ở lần quét sau
                    id_da_doc_cu.update(id_moi_doc)
                    
            except Exception as e:
                them_log(f"Lỗi khi quét: {str(e)}")
                
    except imaplib.IMAP4.error:
        them_log("LỖI: Sai Gmail hoặc Mật khẩu ứng dụng!")
        messagebox.showerror("Lỗi xác thực", "Đăng nhập thất bại! Vui lòng kiểm tra lại mật khẩu ứng dụng.")
        dung_theo_doi()
    except Exception as e:
        them_log(f"Lỗi kết nối: {str(e)}")
        dung_theo_doi()
    finally:
        if mail:
            try:
                mail.logout()
            except:
                pass


def bat_dau_theo_doi():
    global is_monitoring, monitor_thread
    user_email = entry_email.get().strip()
    app_password = entry_pass.get().strip().replace(" ", "")
    
    if not user_email or not app_password:
        messagebox.showwarning("Cảnh báo", "Vui lòng nhập đầy đủ Gmail và Mật khẩu ứng dụng!")
        return

    # Khóa các ô nhập và nút bấm
    entry_email.config(state="disabled")
    entry_pass.config(state="disabled")
    btn_start.config(state="disabled", bg="#9e9e9e")
    btn_stop.config(state="normal", bg="#c62828")
    
    is_monitoring = True
    them_log("Đang kết nối tới máy chủ Google...")
    
    # Khởi tạo luồng ngầm để Tkinter không bị đơ
    monitor_thread = threading.Thread(target=luong_theo_doi_ngam, args=(user_email, app_password), daemon=True)
    monitor_thread.start()


def dung_theo_doi():
    global is_monitoring
    is_monitoring = False
    
    # Mở khóa lại giao diện
    entry_email.config(state="normal")
    entry_pass.config(state="normal")
    btn_start.config(state="normal", bg="#2e7d32")
    btn_stop.config(state="disabled", bg="#9e9e9e")
    
    them_log("Đã dừng theo dõi.")


def khi_dong_ung_dung():
    """Xử lý khi người dùng bấm dấu X tắt ứng dụng"""
    global is_monitoring
    is_monitoring = False
    root.destroy()


# --- Xây dựng giao diện đồ họa Tkinter ---
root = tk.Tk()
root.title("Hệ Thống Theo Dõi Email Tự Động - Hải Đăng Nhóm 11")
root.geometry("500x480")
root.resizable(False, False)
root.configure(bg="#f5f5f5")
root.protocol("WM_DELETE_WINDOW", khi_dong_ung_dung)

tk.Label(
    root, 
    text="THEO DÕI EMAIL ĐÃ ĐỌC TỰ ĐỘNG", 
    font=("Arial", 14, "bold"), 
    fg="#1a237e", 
    bg="#f5f5f5"
).pack(pady=10)

# Khung đăng nhập
frame_login = tk.LabelFrame(root, text=" Xác thực bảo mật Google (2FA) ", font=("Arial", 10, "bold"), padx=15, pady=10, bg="#f5f5f5")
frame_login.pack(fill="x", padx=20, pady=5)

tk.Label(frame_login, text="Tài khoản Gmail:", font=("Arial", 10), bg="#f5f5f5").grid(row=0, column=0, sticky="w", pady=5)
entry_email = tk.Entry(frame_login, font=("Arial", 10), width=30)
entry_email.grid(row=0, column=1, pady=5, padx=10)

tk.Label(frame_login, text="Mật khẩu ứng dụng:", font=("Arial", 10), bg="#f5f5f5").grid(row=1, column=0, sticky="w", pady=5)
entry_pass = tk.Entry(frame_login, font=("Arial", 10), width=30, show="*")
entry_pass.grid(row=1, column=1, pady=5, padx=10)

# Khung nút bấm điều khiển
frame_btn = tk.Frame(root, bg="#f5f5f5")
frame_btn.pack(fill="x", padx=20, pady=10)

btn_start = tk.Button(
    frame_btn, text="BẮT ĐẦU THEO DÕI", bg="#2e7d32", fg="white", font=("Arial", 10, "bold"), 
    width=22, command=bat_dau_theo_doi, relief="raised", cursor="hand2"
)
btn_start.pack(side="left", expand=True, padx=5)

btn_stop = tk.Button(
    frame_btn, text="DỪNG THEO DÕI", bg="#9e9e9e", fg="white", font=("Arial", 10, "bold"), 
    width=22, command=dung_theo_doi, state="disabled", relief="raised", cursor="hand2"
)
btn_stop.pack(side="right", expand=True, padx=5)

# Khung nhật ký hoạt động (Log)
frame_log = tk.LabelFrame(root, text=" Nhật ký hoạt động ", font=("Arial", 10, "bold"), padx=10, pady=5, bg="#f5f5f5")
frame_log.pack(fill="both", expand=True, padx=20, pady=5)

txt_log = scrolledtext.ScrolledText(frame_log, height=10, font=("Consolas", 9), state="disabled", bg="#ffffff")
txt_log.pack(fill="both", expand=True)

tk.Label(
    root, text="Chương trình được làm bởi Hải Đăng Nhóm 11", font=("Arial", 9, "italic"), fg="#555555", bg="#f5f5f5"
).pack(side="bottom", pady=8)

root.mainloop()